import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from streamlit_webrtc import webrtc_streamer

# Config
EXCEL_FILE = 'inventory.xlsx'
SHEET_NAME = 'Sheet1'

# Create dummy Excel if not exists
if not os.path.exists(EXCEL_FILE):
    dummy_data = {
        "Serial Number": ["SN001", "SN002", "SN003", "SN004"],
        "Product Name": ["Widget A", "Widget B", "Widget C", "Widget D"],
        "Quantity": [10, 20, 15, 5],
        "Status": ["Pending", "Pending", "Confirmed", "Pending"]
    }
    dummy_df = pd.DataFrame(dummy_data)
    dummy_df.to_excel(EXCEL_FILE, index=False)

@st.cache_data
def load_excel():
    return pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)

def save_excel(df, color_map):
    df.to_excel(EXCEL_FILE, sheet_name=SHEET_NAME, index=False)
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    for row_idx, color in color_map.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in ws[row_idx + 2]:
            cell.fill = fill
    wb.save(EXCEL_FILE)

# Page config
st.set_page_config(page_title="Blank Inventory Management System", layout="wide")

# Title with white color (uses markdown + HTML)
st.markdown(
    "<h1 style='color: white; background-color: black; padding: 10px;'>Blank Inventory Management System</h1>",
    unsafe_allow_html=True
)

# Load Excel data
df = load_excel()

# Camera section
st.subheader("📸 Capture Image (optional)")
webrtc_streamer(key="camera")

# Search Section
st.subheader("🔍 Search Inventory by Serial Number")
serial_number = st.text_input("Enter Serial Number")
if serial_number:
    matched_rows = df[df['Serial Number'].astype(str) == serial_number]
    if not matched_rows.empty:
        st.success("Match found:")
        st.dataframe(matched_rows)
    else:
        st.warning("No matching serial number found.")

# Manage inventory
st.subheader("✏️ Add / Update / Confirm Inventory")
action = st.selectbox("Select Action", ["Add", "Update", "Confirm"])
row_index = st.number_input("Select Row Index (0-based)", min_value=0, max_value=len(df)-1, step=1)
columns = df.columns.tolist()

st.write("🔧 Edit Row Values:")
user_inputs = {}
for col in columns:
    user_inputs[col] = st.text_input(f"{col}", value=str(df.at[row_index, col]))

if st.button("Submit Action"):
    for col, val in user_inputs.items():
        df.at[row_index, col] = val
    color_map = {}
    if action in ["Add", "Update"]:
        color_map[row_index] = "ADD8E6"  # Light Blue
    elif action == "Confirm":
        color_map[row_index] = "90EE90"  # Light Green
    save_excel(df, color_map)
    st.success(f"{action} completed for row {row_index}")

# Display updated data
st.subheader("📄 Updated Inventory Preview")
st.dataframe(df)

buffer = BytesIO()
df.to_excel(buffer, index=False)
st.download_button(
    "📥 Download Updated Excel File",
    data=buffer.getvalue(),
    file_name="updated_inventory.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
