import streamlit as st
import pandas as pd
from io import BytesIO
import os
import openpyxl

st.set_page_config("Inventory Manager", layout="wide")
st.title("📦 Inventory Manager with Confirmation & Edits")

def update_confirmed_column_based_on_color(wb):
    ws = wb.active

    # Get headers and find/create 'CONFIRMED' column
    headers = [cell.value for cell in ws[1]]
    if "CONFIRMED" not in headers:
        confirmed_col_index = len(headers) + 1
        ws.cell(row=1, column=confirmed_col_index).value = "CONFIRMED"
    else:
        confirmed_col_index = headers.index("CONFIRMED") + 1

    green_rgb_codes = {
        "FFC6EFCE",  # Light green from Excel (common highlight)
        "FF00FF00",  # Standard Excel green
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        first_cell = row[0]
        fill = first_cell.fill
        rgb = ""
        if fill and fill.start_color and fill.start_color.type == "rgb":
            rgb = fill.start_color.rgb

        is_green = rgb in green_rgb_codes
        ws.cell(row=first_cell.row, column=confirmed_col_index).value = "Y" if is_green else "N"

    return wb

uploaded_file = st.file_uploader("Upload your inventory Excel file", type=["xlsx"])

if uploaded_file:
    original_filename = os.path.splitext(uploaded_file.name)[0]

    # Load workbook to check colors
    wb_orig = openpyxl.load_workbook(uploaded_file)
    wb_orig = update_confirmed_column_based_on_color(wb_orig)

    # Save updated original file
    orig_output = BytesIO()
    wb_orig.save(orig_output)
    orig_output.seek(0)

    st.download_button(
        "💾 Download Original Sheet (with CONFIRMED auto-set from green rows)",
        orig_output,
        file_name=f"{original_filename}_with_confirmed.xlsx"
    )

    # Load updated file as dataframe for rest of app
    df = pd.read_excel(orig_output)
    df = df.fillna("")

    st.session_state.setdefault("confirmed_rows", set())
    st.session_state.setdefault("new_rows", set())
    st.session_state.setdefault("edited_cells", {})

    st.subheader("🔍 Search for Machine")
    col_to_search = st.selectbox("Select column to search", df.columns)
    val_to_search = st.text_input("Enter value to search")
    search_btn = st.button("🔍 Search")

    match_indices = []
    match_df = pd.DataFrame()

    if search_btn and val_to_search:
        match_df = df[df[col_to_search].astype(str).str.lower() == val_to_search.lower()]
        match_indices = match_df.index.tolist()
        if match_df.empty:
            st.warning("No match found.")
        else:
            st.success(f"Found {len(match_df)} match(es)")
            st.dataframe(match_df)

    if match_indices:
        selected_index = match_indices[0]
        confirm_btn = st.button("✅ Confirm Match")

        if confirm_btn:
            st.session_state.confirmed_rows.add(selected_index)
            df.at[selected_index, "CONFIRMED"] = "Y"
            st.success("Confirmed. Row will be marked green in final sheet.")

        # Edit Row
        st.subheader("✏️ Edit This Equipment")
        edited = False
        for col in df.columns:
            if col == "CONFIRMED":
                continue
            new_val = st.text_input(f"{col}", value=str(df.at[selected_index, col]), key=f"{col}_edit")
            if new_val != str(df.at[selected_index, col]):
                df.at[selected_index, col] = new_val
                st.session_state.edited_cells[(selected_index, col)] = True
                edited = True

        if edited:
            st.info("Changes saved. Edited cells will be highlighted red.")

    st.subheader("➕ Add New Machine")
    new_data = {}
    for col in df.columns:
        new_data[col] = st.text_input(f"New {col}", key=f"new_{col}")

    if st.button("➕ Add Machine"):
        df.loc[len(df)] = new_data
        df.at[len(df)-1, 'REMARKS'] = "New entry added"
        df.at[len(df)-1, 'CONFIRMED'] = "N"
        st.session_state.new_rows.add(len(df)-1)
        st.success("New machine added.")

    st.subheader("👀 Unmarked Machines")
    unmarked = df[df["CONFIRMED"] != "Y"]
    st.dataframe(unmarked, use_container_width=True)

    # Save Final Colored Sheet
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    df.to_excel(writer, index=False, sheet_name='Inventory')
    workbook = writer.book
    worksheet = writer.sheets['Inventory']

    green_fmt = workbook.add_format({'bg_color': '#C6EFCE'})
    blue_fmt = workbook.add_format({'bg_color': '#ADD8E6'})
    red_fmt = workbook.add_format({'bg_color': '#FFC7CE'})

    for i, row in df.iterrows():
        fmt_row = None
        if i in st.session_state.confirmed_rows:
            fmt_row = green_fmt
        elif i in st.session_state.new_rows:
            fmt_row = blue_fmt

        if fmt_row:
            worksheet.set_row(i + 1, None, fmt_row)

        for j, col in enumerate(df.columns):
            if (i, col) in st.session_state.edited_cells:
                worksheet.write(i + 1, j, row[col], red_fmt)

    writer.close()
    output.seek(0)

    final_name = f"FINAL {original_filename}.xlsx"
    st.download_button("💾 Download FINAL Sheet", output, file_name=final_name)
